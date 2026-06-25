import os

try:
    import openpyxl
except ImportError:
    openpyxl = None

def _check_openpyxl():
    if openpyxl is None:
        raise ImportError("The 'openpyxl' module is required for Excel output. \nInstall it via: pip install openpyxl")

def write_batch(filepath, col_names, data):
    _check_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(col_names)
    for row in data:
        ws.append([row[col] for col in col_names])
    wb.save(filepath)

def initialize_continuous(filepath, col_names):
    _check_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(col_names)
    wb.save(filepath)

def append_continuous(filepath, col_names, row_data):
    _check_openpyxl()
    # Note: Loading and saving full workbooks per row is heavy I/O
    if os.path.exists(filepath):
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(col_names)
    
    ws.append([row_data[col] for col in col_names])
    wb.save(filepath)